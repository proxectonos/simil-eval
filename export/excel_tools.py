import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# Excel Styling Utilities
# ============================================================

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def prettify_excel(excel_path: str):
    """
    Apply clean formatting to an Excel file:
    - Bold blue headers
    - Thin borders and centered alignment
    - Zebra pattern for alternating rows
    - Auto column width adjustment
    - 3-decimal numeric formatting
    - Highlight max numeric value per column (except 'Date')
      * Exception: in surprisal results, if benchmark == 'calame', highlight min instead
    """
    wb = load_workbook(excel_path)
    thin = Side(border_style="thin", color="AAAAAA")

    header_fill = PatternFill("solid", fgColor="DCE6F1")
    zebra_fill = PatternFill("solid", fgColor="F7F7F7")
    highlight_fill = PatternFill("solid", fgColor="C6EFCE")  # light green

    for sheetname in wb.sheetnames:
        ws = wb[sheetname]

        # --- Header formatting ---
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        max_row, max_col = ws.max_row, ws.max_column

        # Find if this Excel likely comes from a surprisal file
        headers = [str(ws.cell(row=1, column=c).value or "").strip().lower() for c in range(1, max_col + 1)]
        is_surprisal = "benchmark" in headers
        benchmark_col_idx = headers.index("benchmark") + 1 if "benchmark" in headers else None

        # --- Base formatting + zebra pattern ---
        for row_idx in range(2, max_row + 1):
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

                # Apply numeric format if applicable
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.000"

                # Apply zebra shading on even rows
                if row_idx % 2 == 0:
                    cell.fill = zebra_fill

        # --- Highlight numeric columns ---
        for col_idx in range(1, max_col + 1):
            header = str(ws.cell(row=1, column=col_idx).value or "").strip().lower()
            if header == "date":
                continue  # skip the Date column

            # Collect numeric values with benchmark info
            values = []
            for row_idx in range(2, max_row + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if isinstance(val, (int, float)):
                    if is_surprisal and benchmark_col_idx:
                        bench = str(ws.cell(row=row_idx, column=benchmark_col_idx).value or "").strip().lower()
                    else:
                        bench = None
                    values.append((row_idx, val, bench))

            if not values:
                continue

            # Decide if we highlight max or min depending on the benchmark
            if is_surprisal and any(v[2] == "calame" for v in values):
                # For Calame benchmark → highlight lowest value
                target_val = min(v[1] for v in values if v[2] == "calame")
                for row_idx, val, bench in values:
                    if bench == "calame" and abs(val - target_val) < 1e-9:
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.fill = highlight_fill
                        cell.font = Font(bold=True)
            else:
                # Default: highlight highest value
                target_val = max(v[1] for v in values)
                for row_idx, val, _ in values:
                    if abs(val - target_val) < 1e-9:
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.fill = highlight_fill
                        cell.font = Font(bold=True)

        # --- Auto column width adjustment ---
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        # --- Freeze header row ---
        ws.freeze_panes = "A2"

    wb.save(excel_path)




# ============================================================
# Generic Helpers
# ============================================================

def collect_results_from_folder(folder_path, parse_function, file_extension=".txt"):
    """
    Generic folder parser: applies a custom parse function to all files with given extension.
    Returns a concatenated DataFrame.
    """
    all_results = []
    for fname in os.listdir(folder_path):
        if not fname.endswith(file_extension):
            continue
        with open(os.path.join(folder_path, fname), encoding="utf-8") as f:
            text = f.read()
        all_results.extend(parse_function(text))
    return pd.DataFrame(all_results)


def save_grouped_by_lang(df, output_path, filename_prefix="results"):
    """
    Save a DataFrame to an Excel file with one sheet per language.
    Returns the full Excel path (caller decides whether to prettify).
    """
    excel_path = os.path.join(output_path, f"{filename_prefix}.xlsx")
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        for lang, group in df.groupby("lang"):
            group_sorted = group.sort_values(by=["model", "benchmark"])
            group_sorted.to_excel(writer, sheet_name=lang.capitalize(), index=False)
    return excel_path
